# ============================================================
# FMCSA MC Number Checker - Output Handler
# ============================================================
# Handles saving results to CSV and Excel formats.
# ============================================================

import os
import csv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# Column headers for output files
HEADERS = [
    "MC Number",
    "Entity Type",
    "USDOT Status",
    "USDOT Number",
    "Operating Authority Status",
    "Legal Name",
    "DBA Name",
    "Physical Address",
    "Phone",
    "Mailing Address",
    "Power Units",
    "Out of Service Date",
    "MCS 150 Form Date",
    "MCS 150 Mileage",
    "Auth For Hire",
]


def _data_to_row(mc_number, data):
    """Convert a parsed data dictionary to an ordered row list."""
    return [
        f"MC-{mc_number}",
        data.get("entity_type", ""),
        data.get("usdot_status", ""),
        data.get("usdot_number", ""),
        data.get("operating_authority_status", ""),
        data.get("legal_name", ""),
        data.get("dba_name", ""),
        data.get("physical_address", ""),
        data.get("phone", ""),
        data.get("mailing_address", ""),
        data.get("power_units", ""),
        data.get("out_of_service_date", ""),
        data.get("mcs150_form_date", ""),
        data.get("mcs150_mileage", ""),
        data.get("auth_for_hire", ""),
    ]


def save_to_csv(results, output_path):
    """
    Save all authorized carrier results to a CSV file.

    Args:
        results: list of tuples (mc_number, data_dict)
        output_path: full path to the output CSV file
    """
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(HEADERS)
        for mc_number, data in results:
            writer.writerow(_data_to_row(mc_number, data))

    print(f"\n  CSV saved: {output_path}")


def save_to_excel(results, output_path):
    """
    Save all authorized carrier results to a styled Excel file.

    Args:
        results: list of tuples (mc_number, data_dict)
        output_path: full path to the output Excel file
    """
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Authorized Carriers"

    # Styling
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    data_font = Font(name="Calibri", size=10)
    data_alignment = Alignment(vertical="center", wrap_text=True)

    alt_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Write headers
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Write data rows
    for row_idx, (mc_number, data) in enumerate(results, 2):
        row_data = _data_to_row(mc_number, data)
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    # Auto fit column widths (approximate)
    column_widths = {
        1: 15,   # MC Number
        2: 15,   # Entity Type
        3: 15,   # USDOT Status
        4: 15,   # USDOT Number
        5: 30,   # Operating Authority Status
        6: 30,   # Legal Name
        7: 20,   # DBA Name
        8: 35,   # Physical Address
        9: 18,   # Phone
        10: 35,  # Mailing Address
        11: 14,  # Power Units
        12: 20,  # Out of Service Date
        13: 20,  # MCS 150 Form Date
        14: 18,  # MCS 150 Mileage
        15: 14,  # Auth For Hire
    }

    for col, width in column_widths.items():
        ws.column_dimensions[chr(64 + col) if col <= 26 else None].width = width

    # Fix column letter for columns > 9
    from openpyxl.utils import get_column_letter
    for col, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    # Add summary sheet
    ws_summary = wb.create_sheet("Summary")
    ws_summary["A1"] = "FMCSA MC Number Scan Report"
    ws_summary["A1"].font = Font(name="Calibri", bold=True, size=14, color="1F4E79")

    ws_summary["A3"] = "Generated:"
    ws_summary["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws_summary["A4"] = "Total Authorized Carriers Found:"
    ws_summary["B4"] = len(results)
    ws_summary["A5"] = "Criteria:"
    ws_summary["B5"] = "Entity Type = CARRIER, USDOT Status = ACTIVE"

    for row in range(3, 6):
        ws_summary.cell(row=row, column=1).font = Font(bold=True)

    ws_summary.column_dimensions["A"].width = 35
    ws_summary.column_dimensions["B"].width = 40

    # Freeze the header row
    ws.freeze_panes = "A2"

    wb.save(output_path)
    print(f"  Excel saved: {output_path}")


def save_progress(mc_number, output_dir, progress_file):
    """
    Save the last checked MC number so scanning can be resumed.
    """
    os.makedirs(output_dir, exist_ok=True)
    progress_path = os.path.join(output_dir, progress_file)
    with open(progress_path, "w") as f:
        f.write(str(mc_number))


def load_progress(output_dir, progress_file):
    """
    Load the last checked MC number from progress file.
    Returns the MC number to resume from, or None.
    """
    progress_path = os.path.join(output_dir, progress_file)
    if os.path.exists(progress_path):
        with open(progress_path, "r") as f:
            content = f.read().strip()
            if content.isdigit():
                return int(content)
    return None


def append_to_csv(mc_number, data, output_path):
    """
    Append a single result row to the CSV file (for real time saving).
    Creates the file with headers if it doesn't exist.
    """
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    file_exists = os.path.exists(output_path)

    with open(output_path, "a", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        if not file_exists:
            writer.writerow(HEADERS)
        writer.writerow(_data_to_row(mc_number, data))
